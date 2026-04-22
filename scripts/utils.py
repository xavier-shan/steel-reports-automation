
from __future__ import annotations

import json
import shutil
from copy import copy
from dataclasses import dataclass
from datetime import datetime, date
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


REQUIRED_PRICE_HEADERS = ["钢厂", "钢种", "贸易商报价", "泉州地区送到价", "涨跌"]
POSSIBLE_HEADER_ALIASES = {
    "钢厂": ["钢厂", "钢厂名称", "厂家", "钢厂/品牌"],
    "钢种": ["钢种", "品种", "规格", "材质"],
    "贸易商报价": ["贸易商报价", "报价", "出厂价", "市场报价", "贸易商报价（三钢只报送到价）"],
    "泉州地区送到价": ["泉州地区送到价", "送到价", "到货价", "送到成本", "泉州地区送到价格"],
    "涨跌": ["涨跌", "今日涨跌"],
    "备注": ["备注", "说明"],
    "和元立差价": ["和元立差价", "与元立差价", "和元立送到差价"],
}


class WorkbookError(RuntimeError):
    pass


def read_json(path: str | Path) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def write_json(path: str | Path, payload: dict) -> None:
    ensure_parent(path)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def ensure_parent(path: str | Path) -> None:
    Path(path).parent.mkdir(parents=True, exist_ok=True)


def iso_today() -> str:
    return date.today().isoformat()


def format_cn_date(input_value: str | None = None) -> str:
    if input_value:
        dt = datetime.fromisoformat(input_value)
    else:
        dt = datetime.now()
    return f"{dt.month}月{dt.day}日"


def safe_filename(prefix: str, date_text: str | None = None, suffix: str = ".xlsx") -> str:
    cn = format_cn_date(date_text)
    return f"{prefix}{cn}{suffix}"


def copy_file(src: str | Path, dst: str | Path) -> None:
    ensure_parent(dst)
    shutil.copy2(src, dst)


def load_xlsx(path: str | Path):
    return load_workbook(path)


def save_xlsx(wb, path: str | Path) -> None:
    ensure_parent(path)
    wb.save(path)


def find_sheet(wb, preferred_name: Optional[str] = None) -> Worksheet:
    if preferred_name and preferred_name in wb.sheetnames:
        return wb[preferred_name]
    return wb[wb.sheetnames[0]]


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def try_float(value: Any) -> Optional[float]:
    if value in ("", None):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = str(value).strip().replace(",", "").replace("，", "")
    if cleaned.endswith(".0") and cleaned[:-2].isdigit():
        return float(cleaned)
    try:
        return float(cleaned)
    except ValueError:
        return None


def clone_cell_style(source, target) -> None:
    if source.has_style:
        target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    if source.font:
        target.font = copy(source.font)
    if source.fill:
        target.fill = copy(source.fill)
    if source.border:
        target.border = copy(source.border)
    if source.alignment:
        target.alignment = copy(source.alignment)
    if source.protection:
        target.protection = copy(source.protection)


@dataclass
class HeaderMap:
    row_index: int
    columns: Dict[str, int]


def _header_matches(cell_value: Any, aliases: Iterable[str]) -> bool:
    text = normalize_text(cell_value)
    return text in aliases


def locate_header_row(
    ws: Worksheet,
    required_headers: Optional[List[str]] = None,
    scan_rows: int = 20,
) -> HeaderMap:
    required_headers = required_headers or REQUIRED_PRICE_HEADERS
    for row_idx in range(1, min(ws.max_row, scan_rows) + 1):
        columns: Dict[str, int] = {}
        for col_idx in range(1, ws.max_column + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            for canonical, aliases in POSSIBLE_HEADER_ALIASES.items():
                if _header_matches(value, aliases):
                    columns[canonical] = col_idx
        if all(h in columns for h in required_headers):
            return HeaderMap(row_index=row_idx, columns=columns)
    raise WorkbookError("未能识别表头，请在配置里手动指定工作表和表头行。")


@dataclass
class SheetRow:
    row_index: int
    mill: str
    grade: str
    trader_price: Optional[float]
    delivered_price: Optional[float]

    def key(self) -> tuple[str, str]:
        return (self.mill, self.grade)


def iter_price_rows(ws: Worksheet, header: HeaderMap) -> List[SheetRow]:
    rows: List[SheetRow] = []
    start = header.row_index + 1
    mill_col = header.columns["钢厂"]
    grade_col = header.columns["钢种"]
    trader_col = header.columns["贸易商报价"]
    delivered_col = header.columns.get("泉州地区送到价")
    last_mill = ""
    for r in range(start, ws.max_row + 1):
        raw_mill = normalize_text(ws.cell(r, mill_col).value)
        grade = normalize_text(ws.cell(r, grade_col).value)
        trader_price = try_float(ws.cell(r, trader_col).value)
        delivered_price = try_float(ws.cell(r, delivered_col).value) if delivered_col else None
        if raw_mill:
            last_mill = raw_mill
        mill = raw_mill or last_mill
        if not grade:
            continue
        if trader_price is None and delivered_price is None:
            continue
        rows.append(
            SheetRow(
                row_index=r,
                mill=mill,
                grade=grade,
                trader_price=trader_price,
                delivered_price=delivered_price,
            )
        )
    return rows


def build_row_index(rows: List[SheetRow]) -> Dict[tuple[str, str], SheetRow]:
    return {row.key(): row for row in rows}


def find_row_by_mill_and_grade(rows: List[SheetRow], mill_names: Iterable[str], grade_names: Iterable[str]) -> Optional[SheetRow]:
    mill_set = set(mill_names)
    grade_set = set(grade_names)
    for row in rows:
        if row.mill in mill_set and row.grade in grade_set:
            return row
    return None


def diff_text(new_value: Optional[float], old_value: Optional[float]) -> str:
    if new_value is None or old_value is None:
        return ""
    delta = round(new_value - old_value)
    if delta > 0:
        return f"涨{abs(delta)}"
    if delta < 0:
        return f"跌{abs(delta)}"
    return "平"


def yuanli_diff_text(current_delivered: Optional[float], yuanli_delivered: Optional[float]) -> str:
    if current_delivered is None or yuanli_delivered is None:
        return ""
    delta = round(current_delivered - yuanli_delivered)
    if delta > 0:
        return f"比元立送到成本高{abs(delta)}"
    if delta < 0:
        return f"比元立送到成本低{abs(delta)}"
    return "和元立送到成本持平"


def set_value(ws: Worksheet, row_idx: int, col_idx: Optional[int], value: Any) -> None:
    if not col_idx:
        return
    ws.cell(row=row_idx, column=col_idx).value = value


def add_row_below(ws: Worksheet, source_row_idx: int) -> int:
    insert_at = source_row_idx + 1
    ws.insert_rows(insert_at, amount=1)
    for col_idx in range(1, ws.max_column + 1):
        clone_cell_style(ws.cell(source_row_idx, col_idx), ws.cell(insert_at, col_idx))
        if ws.row_dimensions[source_row_idx].height:
            ws.row_dimensions[insert_at].height = ws.row_dimensions[source_row_idx].height
    return insert_at
