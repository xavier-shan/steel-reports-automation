"""一次性脚本：生成市场价格日报模板 Excel 文件。"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "工作表1"

# ── 样式定义 ──────────────────────────────────────────
header_font = Font(name="微软雅黑", bold=True, size=11)
header_fill = PatternFill("solid", fgColor="4472C4")
header_font_white = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

title_font = Font(name="微软雅黑", bold=True, size=14)
title_align = Alignment(horizontal="center", vertical="center")

cell_align = Alignment(horizontal="center", vertical="center")
cell_font = Font(name="微软雅黑", size=11)

sangang_fill = PatternFill("solid", fgColor="DEEAF1")
xinxing_fill = PatternFill("solid", fgColor="E2EFDA")
linggang_fill = PatternFill("solid", fgColor="FFF2CC")
xugang_fill  = PatternFill("solid", fgColor="FCE4D6")
yuanli_fill  = PatternFill("solid", fgColor="F4CCFF")

thin = Side(style="thin", color="BFBFBF")
thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_cell(cell, font=None, fill=None, align=None, border=None):
    if font:   cell.font = font
    if fill:   cell.fill = fill
    if align:  cell.alignment = align
    if border: cell.border = border

# ── 列宽 ──────────────────────────────────────────────
col_widths = {1: 20, 2: 12, 3: 16, 4: 18, 5: 10, 6: 22, 7: 45}
for col, width in col_widths.items():
    ws.column_dimensions[get_column_letter(col)].width = width

# ── 第 1 行：标题 ──────────────────────────────────────
ws.row_dimensions[1].height = 36
ws.merge_cells("A1:G1")
title_cell = ws["A1"]
title_cell.value = "市场价格日报"
title_cell.font = Font(name="微软雅黑", bold=True, size=16)
title_cell.alignment = title_align
title_cell.fill = PatternFill("solid", fgColor="1F3864")
title_cell.font = Font(name="微软雅黑", bold=True, size=16, color="FFFFFF")

# ── 第 2 行：表头 ──────────────────────────────────────
ws.row_dimensions[2].height = 30
headers = ["钢厂", "钢种", "贸易商报价", "泉州地区送到价", "涨跌", "和元立差价", "备注"]
for col_idx, h in enumerate(headers, start=1):
    cell = ws.cell(row=2, column=col_idx, value=h)
    style_cell(cell, font=header_font_white, fill=header_fill, align=header_align, border=thin_border)

# ── 数据行 ─────────────────────────────────────────────
# (钢厂, 钢种, 贸易商报价, 泉州地区送到价, 涨跌, 和元立差价, 备注, fill)
rows_data = [
    ("三钢",             "45#",  3700, 3730, "平", "",                        "",  sangang_fill),
    ("三钢",             "40Cr", 3800, 3830, "平", "",                        "",  sangang_fill),
    ("新兴铸管",          "45#",  3680, 3710, "平", "比元立送到成本高150",      "",  xinxing_fill),
    ("凌钢",             "45#",  3690, 3720, "平", "比元立送到成本高160",      "",  linggang_fill),
    ("徐钢，六安，大东海", "45#",  3680, 3700, "平", "比元立送到成本高140",      "",  xugang_fill),
    ("元立",             "45#",  3540, 3560, "平", "",                        "",  yuanli_fill),
    ("元立",             "40Cr", 3640, 3660, "平", "",                        "",  yuanli_fill),
    ("元立",             "08Al", 3460, 3480, "平", "",                        "",  yuanli_fill),
    ("元立",             "195",  3360, 3380, "平", "",                        "",  yuanli_fill),
]

for row_offset, (mill, grade, trader, delivered, change, yuanli_diff, remark, fill) in enumerate(rows_data, start=3):
    ws.row_dimensions[row_offset].height = 22
    values = [mill, grade, trader, delivered, change, yuanli_diff, remark]
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row_offset, column=col_idx, value=val)
        style_cell(cell, font=cell_font, fill=fill, align=cell_align, border=thin_border)

# ── 备注行 ─────────────────────────────────────────────
remark_row = len(rows_data) + 3
ws.row_dimensions[remark_row].height = 22
remark_fill = PatternFill("solid", fgColor="F2F2F2")
ws.merge_cells(f"A{remark_row}:G{remark_row}")
remark_cell = ws[f"A{remark_row}"]
remark_cell.value = "备注：元立建议出厂价45# 3550，以上价格均为含税价格，仅供参考。"
style_cell(remark_cell,
           font=Font(name="微软雅黑", size=10, italic=True, color="595959"),
           fill=remark_fill,
           align=Alignment(horizontal="left", vertical="center", wrap_text=True),
           border=thin_border)

out = "/home/user/steel-reports-automation/template/市场价格日报模板.xlsx"
wb.save(out)
print(f"模板已生成：{out}")
