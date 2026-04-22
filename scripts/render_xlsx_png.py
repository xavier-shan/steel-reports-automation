"""将市场价格日报 Excel 渲染为 PNG 截图。使用 openpyxl 读取数据，matplotlib 绘制表格。"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Any, Optional

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.patches import FancyBboxPatch
import matplotlib.colors as mcolors
from matplotlib import font_manager
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# 注册中文字体
_CJK_FONT_PATHS = [
    "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc",
    "/usr/share/fonts/truetype/wqy/wqy-microhei.ttc",
    "/usr/share/fonts/noto/NotoSansCJK-Regular.ttc",
]
_CJK_FONT_PROP = None
for _fp in _CJK_FONT_PATHS:
    if Path(_fp).exists():
        font_manager.fontManager.addfont(_fp)
        _CJK_FONT_PROP = font_manager.FontProperties(fname=_fp)
        plt.rcParams["font.family"] = _CJK_FONT_PROP.get_name()
        break


def _hex_to_rgb(hex_color: str) -> tuple:
    hex_color = hex_color.lstrip("#")
    if len(hex_color) == 6:
        r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
        return (r / 255, g / 255, b / 255)
    return (1.0, 1.0, 1.0)


def _get_cell_bg(cell) -> Optional[tuple]:
    try:
        fill = cell.fill
        if fill and fill.fill_type == "solid":
            fgc = fill.fgColor
            if fgc.type == "rgb":
                hex_val = fgc.rgb
                if len(hex_val) == 8:
                    hex_val = hex_val[2:]
                if hex_val.upper() not in ("000000", "FFFFFF", "00000000", "FFFFFFFF"):
                    return _hex_to_rgb(hex_val)
    except Exception:
        pass
    return None


def _get_cell_font_color(cell) -> tuple:
    try:
        f = cell.font
        if f and f.color and f.color.type == "rgb":
            hex_val = f.color.rgb
            if len(hex_val) == 8:
                hex_val = hex_val[2:]
            return _hex_to_rgb(hex_val)
    except Exception:
        pass
    return (0.0, 0.0, 0.0)


def _is_bold(cell) -> bool:
    try:
        return bool(cell.font and cell.font.bold)
    except Exception:
        return False


def render(xlsx_path: str, png_path: str) -> None:
    wb = load_workbook(xlsx_path)
    ws = wb.active

    max_row = ws.max_row
    max_col = ws.max_column

    # 收集合并单元格范围
    merged_ranges = {}
    for merged in ws.merged_cells.ranges:
        for r in range(merged.min_row, merged.max_row + 1):
            for c in range(merged.min_col, merged.max_col + 1):
                merged_ranges[(r, c)] = (merged.min_row, merged.min_col,
                                          merged.max_row, merged.max_col)

    # 读取单元格数据
    data = []
    bg_colors = []
    font_colors = []
    bold_flags = []

    for r in range(1, max_row + 1):
        row_data, row_bg, row_fc, row_bold = [], [], [], []
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            val = cell.value
            if val is None:
                # 如果是合并单元格的非主格，取主格的值
                if (r, c) in merged_ranges:
                    mr, mc, _, _ = merged_ranges[(r, c)]
                    if (r, c) != (mr, mc):
                        val = ws.cell(row=mr, column=mc).value
            text = str(val) if val is not None else ""
            row_data.append(text)
            row_bg.append(_get_cell_bg(cell) or (1.0, 1.0, 1.0))
            row_fc.append(_get_cell_font_color(cell))
            row_bold.append(_is_bold(cell))
        data.append(row_data)
        bg_colors.append(row_bg)
        font_colors.append(row_fc)
        bold_flags.append(row_bold)

    # ── 计算列宽（用字符数估计） ────────────────────────────────
    col_weights = []
    for c in range(max_col):
        max_len = max(len(str(data[r][c])) for r in range(max_row))
        col_dim = ws.column_dimensions.get(get_column_letter(c + 1))
        if col_dim and col_dim.width:
            col_weights.append(max(col_dim.width, max_len * 0.8))
        else:
            col_weights.append(max(max_len * 1.2, 8))

    total_weight = sum(col_weights)
    fig_width = min(22, max(14, total_weight * 0.45))
    row_height_pt = 0.45
    fig_height = max_row * row_height_pt + 1.0

    fig, ax = plt.subplots(figsize=(fig_width, fig_height))
    ax.set_xlim(0, fig_width)
    ax.set_ylim(0, max_row * row_height_pt)
    ax.axis("off")

    col_positions = []
    x = 0.0
    for w in col_weights:
        col_positions.append(x)
        x += w / total_weight * fig_width
    col_positions.append(fig_width)

    title_row_drawn = set()

    for r_idx in range(max_row):
        y_top = (max_row - r_idx - 1) * row_height_pt
        row_h = row_height_pt
        try:
            rd = ws.row_dimensions.get(r_idx + 1)
            if rd and rd.height:
                row_h = rd.height / 15.0 * row_height_pt
        except Exception:
            pass

        for c_idx in range(max_col):
            # 合并单元格：跳过非主格的绘制
            key = (r_idx + 1, c_idx + 1)
            is_merged = key in merged_ranges
            if is_merged:
                mr, mc, er, ec = merged_ranges[key]
                if (r_idx + 1, c_idx + 1) != (mr, mc):
                    continue
                # 计算合并后的实际宽度
                x_start = col_positions[mc - 1]
                x_end = col_positions[ec]
                cell_width = x_end - x_start
            else:
                x_start = col_positions[c_idx]
                cell_width = col_positions[c_idx + 1] - x_start

            bg = bg_colors[r_idx][c_idx]
            text = data[r_idx][c_idx]
            fc = font_colors[r_idx][c_idx]
            bold = bold_flags[r_idx][c_idx]

            rect = plt.Rectangle(
                (x_start, y_top), cell_width, row_h,
                facecolor=bg, edgecolor=(0.75, 0.75, 0.75), linewidth=0.4
            )
            ax.add_patch(rect)

            if text:
                fontsize = 8 if r_idx == 0 else 7
                weight = "bold" if bold else "normal"
                ax.text(
                    x_start + cell_width / 2,
                    y_top + row_h / 2,
                    text,
                    ha="center", va="center",
                    fontsize=fontsize,
                    fontweight=weight,
                    color=fc,
                    clip_on=True,
                    wrap=False,
                )

    plt.tight_layout(pad=0.3)
    plt.savefig(png_path, dpi=200, bbox_inches="tight", facecolor="white")
    plt.close()
    print(f"截图已保存：{png_path}")


def main():
    parser = argparse.ArgumentParser(description="Excel → PNG 渲染")
    parser.add_argument("--xlsx", required=True)
    parser.add_argument("--png", required=True)
    args = parser.parse_args()
    render(args.xlsx, args.png)


if __name__ == "__main__":
    main()
